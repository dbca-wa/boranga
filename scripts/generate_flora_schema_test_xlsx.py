"""
Comprehensive test xlsx for Flora Master Schema Whole (id=132).

Populates as many fields as possible across all model groups (ORFCON, ORFSUB,
ORFLOC, ORFGEO, ORFHAB, ORFHQ, ORFVEG, ORFFH, ORFOBS, ORFNUM, ORFID, ORFDOC,
ORFTHR) in addition to exercising all OCC linking paths via OCC + ORFAPP columns.

Run from the project root:
    ./manage.py shell < scripts/generate_flora_schema_test_xlsx.py

Test rows
---------
  Row 2  -- OCC path: create new OCC; all field groups fully populated.
  Row 3  -- OCC path: link existing OCC (same migrated_from_id); OCC fields NOT overwritten.
  Row 4  -- ORFAPP Occurrence fallback: 'mega-flora-occ-001' resolved via migrated_from_id.
  Row 5  -- ORFAPP path: create new OCC via new_occurrence_name + officer.
  Row 6  -- ORFAPP path: link approved ORF to existing OCC via ORFAPP Occurrence FK.
  Row 7  -- with_assessor ORF: OCC columns present but silently ignored (occ=None).
  Row 8  -- No OCC columns: ORF created with no OCC link.
  Row 9  -- UPDATE MODE: add a second document to ORF 001 (only ORFDOC columns + migrated_from_id).
  Row 10 -- UPDATE MODE: add a second threat to ORF 001 (only ORFTHR columns + migrated_from_id).
"""

import json
import os
import sys
from datetime import datetime

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "boranga.settings")

import openpyxl  # noqa: E402
from django.conf import settings  # noqa: E402
from django.contrib.contenttypes.models import ContentType  # noqa: E402
from ledger_api_client.ledger_models import EmailUserRO  # noqa: E402
from openpyxl.styles import Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from boranga.components.occurrence.models import (  # noqa: E402
    Occurrence,
    OccurrenceReportApprovalDetails,
    OccurrenceReportBulkImportSchema,
    OccurrenceReportBulkImportSchemaColumn,
    ThreatAgent,
    ThreatCategory,
)
from boranga.components.species_and_communities.models import Species as _SpeciesModel  # noqa: E402
from boranga.components.users.models import SystemGroup  # noqa: E402

# Unique string included in row data so each test run produces distinct row hashes
# and is not rejected by the duplicate-detection check.
RUN_TS = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

SCHEMA_ID = 132

# ---------------------------------------------------------------------------
# 1. Load schema
# ---------------------------------------------------------------------------
schema = OccurrenceReportBulkImportSchema.objects.get(id=SCHEMA_ID)
flora_gt = schema.group_type
print(f"Schema: {schema.name!r} (id={schema.id}, group_type={flora_gt.name})")
print(f"Existing column count: {schema.columns.count()}")


# ---------------------------------------------------------------------------
# 2. Assessor / approver and officer user ID
# ---------------------------------------------------------------------------
def _first_group_member(group_name):
    sg = SystemGroup.objects.filter(name=group_name).first()
    if not sg:
        return None, None
    for uid in sg.get_system_group_member_ids():
        u = EmailUserRO.objects.filter(id=uid).first()
        if u and u.email:
            return u.email, u.id
    return None, None


assessor_email, assessor_id = _first_group_member(settings.GROUP_NAME_OCCURRENCE_ASSESSOR)
approver_email, _ = _first_group_member(settings.GROUP_NAME_OCCURRENCE_APPROVER)
if not assessor_email or not approver_email:
    print("ERROR: Could not find assessor/approver emails.")
    sys.exit(1)
print(f"Assessor: {assessor_email} (id={assessor_id})")
print(f"Approver: {approver_email}")

# ---------------------------------------------------------------------------
# 3. Flora species and existing OCC for linking tests
# ---------------------------------------------------------------------------
_flora_species = (
    _SpeciesModel.objects.filter(group_type=flora_gt)
    .exclude(taxonomy__scientific_name="")
    .select_related("taxonomy")
    .first()
)
if not _flora_species:
    print("ERROR: No flora species with a scientific name found.")
    sys.exit(1)
flora_species_lookup = _flora_species.taxonomy.scientific_name
print(f"Species: {_flora_species.species_number!r} ({flora_species_lookup!r})")

existing_occ = Occurrence.objects.filter(group_type=flora_gt).order_by("-id").first()
if not existing_occ:
    print("ERROR: No existing flora OCC found.")
    sys.exit(1)
print(f"Existing OCC for linking: {existing_occ.occurrence_number}")

# ---------------------------------------------------------------------------
# 4. Runtime FK lookups for large tables
# ---------------------------------------------------------------------------
_threat_cat = ThreatCategory.objects.filter(archived=False).first()
threat_cat_name = _threat_cat.name if _threat_cat else None
print(f"Threat category: {threat_cat_name!r}")

_threat_agent = ThreatAgent.objects.filter(archived=False).first()
threat_agent_name = _threat_agent.name if _threat_agent else None
print(f"Threat agent: {threat_agent_name!r}")

# ---------------------------------------------------------------------------
# 5. Add missing OCC / ORFAPP columns (idempotent)
# ---------------------------------------------------------------------------
occ_ct = ContentType.objects.get_for_model(Occurrence)
orfapp_ct = ContentType.objects.get_for_model(OccurrenceReportApprovalDetails)

existing_occ_field_names = set(
    schema.columns.filter(django_import_content_type=occ_ct).values_list("django_import_field_name", flat=True)
)
existing_orfapp_field_names = set(
    schema.columns.filter(django_import_content_type=orfapp_ct).values_list("django_import_field_name", flat=True)
)
print(f"\nExisting OCC fields in schema: {sorted(existing_occ_field_names)}")
print(f"Existing ORFAPP fields in schema: {sorted(existing_orfapp_field_names)}")

NEW_COLUMNS = [
    (occ_ct, "OCC Wild Status", "wild_status", True, False),
    (occ_ct, "OCC Comment", "comment", True, False),
    (occ_ct, "OCC Processing Status", "processing_status", True, False),
    (orfapp_ct, "ORFAPP New Occurrence Name", "new_occurrence_name", True, False),
    (orfapp_ct, "ORFAPP Officer", "officer", True, False),
    (orfapp_ct, "ORFAPP Occurrence", "occurrence", True, False),
]

added = 0
for ct, header, field, allow_blank, is_emailuser in NEW_COLUMNS:
    existing = existing_occ_field_names if ct == occ_ct else existing_orfapp_field_names
    if field in existing:
        print(f"  SKIP (already exists): {header}")
        continue
    col = OccurrenceReportBulkImportSchemaColumn(
        schema=schema,
        xlsx_column_header_name=header,
        django_import_content_type=ct,
        django_import_field_name=field,
        xlsx_data_validation_allow_blank=allow_blank,
        is_emailuser_column=is_emailuser,
    )
    col.save()
    print(f"  ADDED: {header}")
    added += 1

print(f"\nAdded {added} new column(s). Schema now has {schema.columns.count()} columns.")

# ---------------------------------------------------------------------------
# 6. Build column header list (in schema order)
# ---------------------------------------------------------------------------
headers = list(schema.columns.order_by("order").values_list("xlsx_column_header_name", flat=True))
print(f"Headers ({len(headers)}): last 12 = {headers[-12:]}")

# ---------------------------------------------------------------------------
# 7. GeoJSON geometry (Perth CBD point)
# ---------------------------------------------------------------------------
PERTH_POINT = json.dumps(
    {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [115.8605, -31.9505]},
                "properties": {},
            }
        ],
    }
)

# ---------------------------------------------------------------------------
# 8. Rich shared field data applied to approved rows
# ---------------------------------------------------------------------------
# Value format notes:
#   FK field (any size): pass the display name (looked up via model 'name' field).
#   MultiSelectField (land_form, soil_type): comma-separated display names.
#   CharField with choices (count_status, OCC processing_status): pass the KEY.
#   BooleanField: True / False (Python bool, written as xlsx boolean).
#   IntegerField: Python int.
#   DecimalField: Python float (validator converts via Decimal()).
#   DateField: "DD/MM/YYYY" string (parsed by validator).

RICH_FIELDS = {
    # Observer contact details
    "ORFCON Observer Name": "A. Botanist",
    "ORFCON Role": "Botanist",
    "ORFCON Contact": "a.botanist@dbca.wa.gov.au",
    "ORFCON Organisation": "DBCA",
    "ORFCON Main Observer": True,
    # Submitter category
    "ORFSUB Submitter Category": "DBCA",
    # Location
    "ORFLOC Location Description": "Open Wandoo woodland on undulating sandplain.",
    "ORFLOC Boundary Description": "Bounded by cleared farmland (N) and road reserve (E).",
    "ORFLOC Mapped Boundary": True,
    "ORFLOC Coordinate Source": "GPS",
    "ORFLOC Location Accuracy": "~300m",
    "ORFLOC Region": "South West",
    "ORFLOC District": "Swan Coastal",
    "ORFLOC Locality": "Test Locality WA 6000",
    # Geometry
    "ORFGEO Geometry": PERTH_POINT,
    # Habitat composition
    "ORFHAB Land Form": "Sandplain",  # MultiSelectField display name
    "ORFHAB Rock Type": "Granite",
    "ORFHAB Loose Rock Percent": 5,  # IntegerField
    "ORFHAB Soil Type": "Sand",  # MultiSelectField display name
    "ORFHAB Soil Colour": "Red",
    "ORFHAB Soil Condition": "Dry",
    "ORFHAB Drainage": "Well drained",
    "ORFHAB Water Quality": "Brackish",  # free-text CharField
    "ORFHAB Habitat Notes": f"Open Wandoo woodland with sparse Acacia understorey. [{RUN_TS}]",
    # Habitat condition -- DecimalField percentages
    "ORFHQ Pristine": 10.0,
    "ORFHQ Excellent": 20.0,
    "ORFHQ Very Good": 30.0,
    "ORFHQ Good": 25.0,
    "ORFHQ Degraded": 10.0,
    "ORFHQ Completely Degraded": 5.0,
    "ORFHQ Count Date": "01/05/2027",  # DateField
    # Vegetation structure -- free-text TextField
    "ORFVEG Vegetation Structure Layer One": "Eucalyptus wandoo (Wandoo) canopy 8-12m",
    "ORFVEG Vegetation Structure Layer Two": "Acacia pulchella (Prickly Moses) 1-2m",
    "ORFVEG Vegetation Structure Layer Three": "Lomandra sp. ground layer",
    "ORFVEG Vegetation Structure Layer Four": "Annual grasses and herbs",
    # Fire history
    "ORFFH Last Fire Estimate": "01/01/2020",  # DateField
    "ORFFH Intensity": "Low",
    "ORFFH Comment": "Low intensity fire approximately 5 years ago.",
    # Observation detail
    "ORFOBS Observation Method": "Opportunistic",
    "ORFOBS Area Surveyed": 500.0,  # DecimalField (m2)
    "ORFOBS Survey Duration": 2,  # IntegerField (hours)
    # Plant count -- flora-specific ORFNUM
    "ORFNUM Plant Count Method": "Actual count - individuals",
    "ORFNUM Plant Count Accuracy": "Actual",
    "ORFNUM Counted Subject": "Plants",
    "ORFNUM Plant Condition": "Healthy",
    "ORFNUM Estimated Population Area": 100.0,
    "ORFNUM Counted": "simple_count",  # CharField key
    "ORFNUM Simple Alive": 150,
    "ORFNUM Simple Dead": 10,
    "ORFNUM Quadrats Present": False,
    "ORFNUM Quadrats Data Attached": False,
    "ORFNUM Quadrats Surveyed": 0,
    "ORFNUM Clonal Reproduction Present": False,
    "ORFNUM Vegetative State Present": True,
    "ORFNUM Flower Bud Present": True,
    "ORFNUM Flower Present": True,
    "ORFNUM Immature Fruit Present": False,
    "ORFNUM Ripe Fruit Present": True,
    "ORFNUM Dehisced Fruit Present": False,
    "ORFNUM Flowering Plants Per": 75.0,  # DecimalField (%)
    "ORFNUM Pollinator Observation": "Native bees observed foraging.",
    # Identification
    "ORFID Id Confirmed By": "Dr Jane Smith",
    "ORFID Identification Certainty": "High Certainty",
    "ORFID Sample Type": "Spirit specimen",
    "ORFID Sample Destination": "WA Herbarium",
    "ORFID Permit Type": "Flora Type 1",
    "ORFID Permit Id": "FL-2027-001",
    "ORFID Collector Number": "COL-001",
    "ORFID Barcode Number": "BAR-001",
    "ORFID Identification Comment": "Confirmed by WA Herbarium specimen comparison.",
    # Document (_file left blank -- no attachment zip provided)
    "ORFDOC Name": "Test Survey Report 2027",
    "ORFDOC Description": "Survey report for comprehensive import testing.",
    "ORFDOC Can Submitter Access": True,
    "ORFDOC Document Category": "ORF Document",
    "ORFDOC Document Sub Category": "Survey Report",
    # Threat
    "ORFTHR Threat Category": threat_cat_name,
    "ORFTHR Threat Agent": threat_agent_name,
    "ORFTHR Current Impact": "Low",
    "ORFTHR Potential Impact": "Low",
    "ORFTHR Potential Threat Onset": "Short Term (under 1 yr)",
    "ORFTHR Date Observed": "01/05/2027",
}

_approved = {
    "ORF Assigned Officer": assessor_email,
    "ORF Assigned Approver": approver_email,
}

# ---------------------------------------------------------------------------
# 9. Generate xlsx
# ---------------------------------------------------------------------------
OUT_PATH = os.path.join(settings.BASE_DIR, "flora_schema_test.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Import"

ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)


def make_row(**kwargs):
    row = {h: "" for h in headers}
    row.update(kwargs)
    return [row[h] for h in headers]


# -- Row 2: OCC path -- create new OCC, all field groups populated --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-001",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "01/05/2027",
            "ORF Assessor Data": f"Assessor notes for row 2 [{RUN_TS}]",
            "ORF Site": "Test Site Alpha",
            "OCC Migrated From Id": "mega-flora-occ-001",
            "OCC Processing Status": "active",
            "OCC Wild Status": "Flora Translocation",
            "OCC Comment": "Created by flora comprehensive test row 2",
            "ORFAPP New Occurrence Name": "Mega Flora OCC 001",
            "ORFAPP Officer": assessor_id,
            **_approved,
            **RICH_FIELDS,
        }
    )
)

# -- Row 3: OCC path -- link same OCC; OCC fields NOT overwritten --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-002",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "15/05/2027",
            "ORF Site": "Test Site Beta",
            "OCC Migrated From Id": "mega-flora-occ-001",
            "OCC Comment": "SHOULD NOT overwrite (existing OCC)",
            "ORFAPP Occurrence": "mega-flora-occ-001",
            "ORFAPP Officer": assessor_id,
            **_approved,
            **RICH_FIELDS,
        }
    )
)

# -- Row 4: ORFAPP fallback -- 'mega-flora-occ-001' resolved via migrated_from_id --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-003",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "20/05/2027",
            "ORFAPP Occurrence": "mega-flora-occ-001",
            "ORFAPP Officer": assessor_id,
            **_approved,
            **RICH_FIELDS,
        }
    )
)

# -- Row 5: ORFAPP path -- create new OCC via new_occurrence_name --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-004",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "01/06/2027",
            "ORFAPP New Occurrence Name": "Mega Flora ORFAPP Created OCC",
            "ORFAPP Officer": assessor_id,
            **_approved,
            **RICH_FIELDS,
        }
    )
)

# -- Row 6: ORFAPP path -- link to existing OCC via occurrence_number FK --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-005",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "10/06/2027",
            "ORFAPP Occurrence": existing_occ.occurrence_number,
            "ORFAPP Officer": assessor_id,
            **_approved,
            **RICH_FIELDS,
        }
    )
)

# -- Row 7: with_assessor ORF -- OCC columns silently ignored --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-006",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "with_assessor",
            "ORF Observation Date": "15/06/2027",
            "ORF Assigned Officer": assessor_email,
            "OCC Migrated From Id": "mega-flora-occ-ignored",
            "OCC Processing Status": "draft",
            "OCC Wild Status": "Flora Translocation",
            "OCC Comment": "these OCC columns should be ignored",
            "ORFCON Observer Name": "C. Observer",
            "ORFCON Role": "Ecologist",
            "ORFCON Organisation": "DBCA",
            "ORFLOC Location Description": f"Location for with_assessor row [{RUN_TS}]",
            "ORFLOC Region": "South West",
            "ORFOBS Observation Method": "Survey (specify type in comments)",
        }
    )
)

# -- Row 8: No OCC fields, minimal data --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-007",
            "ORF Species": flora_species_lookup,
            "ORF Processing Status": "with_assessor",
            "ORF Observation Date": "20/06/2027",
            "ORF Assigned Officer": assessor_email,
            "ORFCON Observer Name": "D. Observer",
            "ORFCON Role": "Conservation officer",
            "ORFLOC Location Description": f"Minimal row location [{RUN_TS}]",
        }
    )
)

# -- Row 9: UPDATE MODE -- add a second document to ORF 001 --
# Only ORF Migrated From ID and ORFDOC columns populated; all others blank.
# The importer detects the same prefixed migrated_from_id as row 2 and switches
# to update mode, creating a new OccurrenceReportDocument on the existing ORF.
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-001",
            "ORFDOC Name": f"Additional Document [{RUN_TS}]",
            "ORFDOC Description": "Second document added via update-mode sub-record row.",
            "ORFDOC Can Submitter Access": True,
            "ORFDOC Document Category": "ORF Document",
            "ORFDOC Document Sub Category": "Survey Report",
        }
    )
)

# -- Row 10: UPDATE MODE -- add a second threat to ORF 001 --
# Same ORF Migrated From ID, only ORFTHR columns populated.
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-flora-orf-001",
            "ORFTHR Threat Category": threat_cat_name,
            "ORFTHR Threat Agent": threat_agent_name,
            "ORFTHR Current Impact": "Medium",
            "ORFTHR Potential Impact": "High",
            "ORFTHR Potential Threat Onset": "Short Term (under 1 yr)",
            "ORFTHR Date Observed": "01/05/2027",
        }
    )
)

# Auto-width columns
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value))) + 4
for col, w in dims.items():
    ws.column_dimensions[get_column_letter(col)].width = min(w, 40)

wb.save(OUT_PATH)
print(f"\nTest xlsx written to: {OUT_PATH}")
print(f"Schema id: {schema.id}")
print(f"Existing OCC for ORFAPP FK row: {existing_occ.occurrence_number}")
print("""
Expected behaviour
------------------
Row 2: Creates OCC (mega-flora-occ-001) named 'Mega Flora OCC 001'; approved.
        ALL field groups populated: ORFCON, ORFSUB, ORFLOC, ORFGEO, ORFHAB,
        ORFHQ, ORFVEG, ORFFH, ORFOBS, ORFNUM, ORFID, ORFDOC, ORFTHR.
Row 3: Links ORF to same OCC; OCC fields NOT overwritten; rich fields saved.
Row 4: ORFAPP Occurrence resolved via migrated_from_id fallback.
Row 5: ORFAPP path creates new OCC 'Mega Flora ORFAPP Created OCC'.
Row 6: ORFAPP path links ORF to existing OCC via occurrence_number FK.
Row 7: with_assessor -- OCC columns silently ignored (occ=None).
Row 8: with_assessor -- no OCC columns, minimal data.
mega-flora-occ-ignored NOT created.
""")
