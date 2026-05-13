"""
Add remaining allowed OCC columns to the 'Fauna Whole Dec 2025' schema (id=137)
and generate a test xlsx that exercises every OCC linking path.

Run from the project root:
    ./manage.py shell < scripts/generate_mega_schema_test_xlsx.py

Occurrence.BULK_IMPORT_INCLUDE_FIELDS allowlist (the only OCC fields valid in schemas):
    wild_status       -- already in schema
    comment           -- already in schema
    migrated_from_id  -- added by this script (idempotent)
    processing_status -- added by this script (idempotent)

Mandatory ORFAPP field rules for approved OCRs:
    officer           -- ALWAYS mandatory
    new_occurrence_name -- mandatory when CREATING a new OCC (any path)
    occurrence        -- mandatory when LINKING an existing OCC (any path)

Test rows
---------
  Row 2  -- OCC path: create new OCC; new_occurrence_name + officer mandatory
  Row 3  -- OCC path: link existing OCC; occurrence (fallback) + officer mandatory
  Row 4  -- ORFAPP occurrence fallback: short migrated_from_id resolved via migrated_from_id
  Row 5  -- ORFAPP path: create new OCC via new_occurrence_name + officer
  Row 6  -- ORFAPP path: link existing OCC via occurrence FK (occurrence_number) + officer
  Row 7  -- with_assessor ORF: OCC columns ignored (status != with_approver/approved)
  Row 8  -- No OCC columns at all: ORF created with no OCC link
"""

import os
import sys

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "boranga.settings")

import openpyxl
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from ledger_api_client.ledger_models import EmailUserRO
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from boranga.components.occurrence.models import (
    Occurrence,
    OccurrenceReportBulkImportSchema,
    OccurrenceReportBulkImportSchemaColumn,
)
from boranga.components.species_and_communities.models import Species as _SpeciesModel
from boranga.components.users.models import SystemGroup

SCHEMA_ID = 137

# ---------------------------------------------------------------------------
# 1. Load schema
# ---------------------------------------------------------------------------
schema = OccurrenceReportBulkImportSchema.objects.get(id=SCHEMA_ID)
fauna_gt = schema.group_type
print(f"Schema: {schema.name!r} (id={schema.id}, group_type={fauna_gt.name})")
print(f"Existing column count: {schema.columns.count()}")


# ---------------------------------------------------------------------------
# 2. Assessor / approver emails
# ---------------------------------------------------------------------------
def _first_group_member_email(group_name):
    sg = SystemGroup.objects.filter(name=group_name).first()
    if not sg:
        return None
    for uid in sg.get_system_group_member_ids():
        u = EmailUserRO.objects.filter(id=uid).first()
        if u and u.email:
            return u.email
    return None


assessor_email = _first_group_member_email(settings.GROUP_NAME_OCCURRENCE_ASSESSOR)
approver_email = _first_group_member_email(settings.GROUP_NAME_OCCURRENCE_APPROVER)
if not assessor_email or not approver_email:
    print("ERROR: Could not find assessor/approver emails.")
    sys.exit(1)
print(f"Assessor: {assessor_email}")
print(f"Approver: {approver_email}")

# ---------------------------------------------------------------------------
# 3. Fauna species and existing OCC for linking tests
# ---------------------------------------------------------------------------
# ORF Species uses species_number as lookup field

_fauna_species = _SpeciesModel.objects.filter(group_type=fauna_gt).exclude(species_number="").first()
if not _fauna_species:
    print("ERROR: No fauna species found.")
    sys.exit(1)

fauna_species_number = _fauna_species.species_number
print(f"Species: {fauna_species_number!r}")

existing_occ = (
    Occurrence.objects.filter(occurrence_number="OCC272230").first()
    or Occurrence.objects.filter(group_type=fauna_gt).order_by("-id").first()
)
if not existing_occ:
    print("ERROR: No existing fauna OCC found.")
    sys.exit(1)
print(f"Existing OCC for linking: {existing_occ.occurrence_number}")

# ---------------------------------------------------------------------------
# 4. Add missing OCC columns (idempotent)
# ---------------------------------------------------------------------------
occ_ct = ContentType.objects.get_for_model(Occurrence)

existing_occ_field_names = set(
    schema.columns.filter(django_import_content_type=occ_ct).values_list("django_import_field_name", flat=True)
)
print(f"\nExisting OCC fields in schema: {sorted(existing_occ_field_names)}")

# New OCC columns to add if not present.
# Only fields in Occurrence.BULK_IMPORT_INCLUDE_FIELDS are valid schema columns.
# wild_status and comment are already present; add migrated_from_id and processing_status.
# Tuples: (header, field_name, allow_blank)
NEW_OCC_COLUMNS = [
    ("OCC Migrated From ID", "migrated_from_id", True),
    ("OCC Processing Status", "processing_status", True),
]

added = 0
for header, field, allow_blank in NEW_OCC_COLUMNS:
    if field in existing_occ_field_names:
        print(f"  SKIP (already exists): {header}")
        continue
    col = OccurrenceReportBulkImportSchemaColumn(
        schema=schema,
        xlsx_column_header_name=header,
        django_import_content_type=occ_ct,
        django_import_field_name=field,
        xlsx_data_validation_allow_blank=allow_blank,
        is_emailuser_column=False,
    )
    col.save()
    print(f"  ADDED: {header}")
    added += 1

print(f"\nAdded {added} new OCC column(s). Schema now has {schema.columns.count()} columns.")

# ---------------------------------------------------------------------------
# 5. Build column header list (in schema order)
# ---------------------------------------------------------------------------
headers = list(schema.columns.order_by("order").values_list("xlsx_column_header_name", flat=True))
print(f"Headers ({len(headers)}): {headers[-10:]}")  # show last 10 to confirm new cols

# ---------------------------------------------------------------------------
# 6. Generate xlsx
# ---------------------------------------------------------------------------
OUT_PATH = os.path.join(settings.BASE_DIR, "mega_schema_test.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Import"

ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)


def make_row(**kwargs):
    row = {h: "" for h in headers}
    row.update(kwargs)
    return [row[h] for h in headers]


_approved = {
    "ORF Assigned Officer": assessor_email,
    "ORF Assigned Approver": approver_email,
    "ORF Approved By": approver_email,
}

# -- Row 2: OCC path, create new OCC with all allowed OCC fields --
# new_occurrence_name + officer are mandatory even when using OCC columns to create.
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-001",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "01/03/2026",
            "ORFAPP New Occurrence Name": "Mega OCC 001",
            "ORFAPP Officer": 109,
            "OCC Migrated From ID": "mega-occ-001",
            "OCC Processing Status": "active",
            "OCC Wild Status": "Fauna Translocation",
            "OCC Comment": "Created by mega test row 2",
            **_approved,
        }
    )
)

# -- Row 3: OCC path, same migrated_from_id -- links to same OCC, fields NOT overwritten --
# occurrence + officer are mandatory when linking via OCC columns to an existing OCC.
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-002",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "15/03/2026",
            "ORFAPP Occurrence": "mega-occ-001",  # fallback via migrated_from_id
            "ORFAPP Officer": 109,
            "OCC Migrated From ID": "mega-occ-001",
            "OCC Comment": "SHOULD NOT overwrite (existing OCC)",
            **_approved,
        }
    )
)

# -- Row 4: ORFAPP fallback test: ORFAPP Occurrence = short migrated_from_id of OCC from row 2 --
# occurrence_number lookup fails → fallback tries migrated_from_id → succeeds
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-003",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "20/03/2026",
            "ORFAPP Occurrence": "mega-occ-001",
            "ORFAPP Officer": 109,
            **_approved,
        }
    )
)

# -- Row 5: ORFAPP path, create new OCC via new_occurrence_name --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-004",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "01/04/2026",
            "ORFAPP New Occurrence Name": "Mega ORFAPP Created OCC",
            "ORFAPP Officer": 109,
            **_approved,
        }
    )
)

# -- Row 6: ORFAPP path, link to existing OCC via ORFAPP Occurrence FK --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-005",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "10/04/2026",
            "ORFAPP Occurrence": existing_occ.occurrence_number,
            "ORFAPP Officer": 109,
            **_approved,
        }
    )
)

# -- Row 7: with_assessor ORF, OCC columns ignored because status != with_approver/approved --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-006",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "with_assessor",
            "ORF Observation Date": "15/04/2026",
            "ORF Assigned Officer": assessor_email,
            "OCC Migrated From ID": "mega-occ-ignored",
            "OCC Processing Status": "draft",
            "OCC Wild Status": "Fauna Translocation",
            "OCC Comment": "these OCC columns should be ignored",
        }
    )
)

# -- Row 8: No OCC fields at all --
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "mega-orf-007",
            "ORF Species": fauna_species_number,
            "ORF Processing Status": "with_assessor",
            "ORF Observation Date": "20/04/2026",
            "ORF Assigned Officer": assessor_email,
        }
    )
)

# Auto-width
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value))) + 4
for col, w in dims.items():
    ws.column_dimensions[get_column_letter(col)].width = min(w, 40)

wb.save(OUT_PATH)
print(f"\nTest xlsx written to: {OUT_PATH}")
print(f"Schema id: {schema.id}")
print(f"Existing OCC for ORFAPP FK row: {existing_occ.occurrence_number}")
print("""
Expected behaviour
------------------
Row 2: Creates OCC (mega-occ-001) named 'Mega OCC 001' via OCC columns + ORFAPP new_occurrence_name;
        approved ORF linked. OCC fields (active, Fauna Translocation, comment) applied.
Row 3: Links second approved ORF to same OCC via OCC migrated_from_id + ORFAPP Occurrence fallback;
        OCC fields NOT re-applied (existing OCC).
Row 4: ORFAPP Occurrence = 'mega-occ-001' (fallback via migrated_from_id) + officer mandatory.
Row 5: ORFAPP path -- creates new OCC 'Mega ORFAPP Created OCC'; ORF linked.
Row 6: ORFAPP path -- links approved ORF to existing OCC via ORFAPP Occurrence FK.
Row 7: with_assessor ORF: OCC columns present but silently ignored (occ=None).
Row 8: No OCC columns -- ORF created, no OCC created or linked.
""")
