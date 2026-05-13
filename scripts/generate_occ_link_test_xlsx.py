"""
Generate a bulk-import xlsx to test OCC-linking via both OCC columns and
OccurrenceReportApprovalDetails (ORFAPP) columns, including ambiguity detection.

Run from the project root:
    ./manage.py shell < scripts/generate_occ_link_test_xlsx.py

Test rows
---------
  Row 2  - OCC path: creates OCC "occ-test-001", approved ORF linked via OCC Migrated From ID
  Row 3  - OCC path: second approved ORF links to same OCC via same migrated_from_id
  Row 4  - OCC path: with_assessor ORF; OCC linked (occurrence FK set via reverse-FK add)
  Row 5  - OCC path: with_assessor ORF; no OCC fields -> no OCC created
  Row 6  - ORFAPP path: approved ORF creates a new OCC via ORFAPP New Occurrence Name
  Row 7  - ORFAPP path: approved ORF links to existing OCC via ORFAPP Occurrence FK
  Row 8  - AMBIGUOUS: approved ORF with BOTH OCC Migrated From ID AND ORFAPP New Occurrence Name -> error
  Row 9  - AMBIGUOUS: approved ORF with BOTH OCC Migrated From ID AND ORFAPP Occurrence FK -> error
  Row 10 - AMBIGUOUS: approved ORF with BOTH ORFAPP Occurrence FK AND ORFAPP New Occurrence Name -> error
"""

import os
import sys

import openpyxl

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "boranga.settings")

from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from ledger_api_client.ledger_models import EmailUserRO
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from boranga.components.occurrence.models import (
    GroupType,
    Occurrence,
    OccurrenceReport,
    OccurrenceReportApprovalDetails,
    OccurrenceReportBulkImportSchema,
    OccurrenceReportBulkImportSchemaColumn,
)
from boranga.components.species_and_communities.models import Taxonomy
from boranga.components.users.models import SystemGroup

# ---------------------------------------------------------------------------
# 1. Fauna group type
# ---------------------------------------------------------------------------
fauna_gt = GroupType.objects.get(name="fauna")
print(f"Fauna group type id: {fauna_gt.id}")

# ---------------------------------------------------------------------------
# 2. Real assessor/approver emails (required for approved rows)
# ---------------------------------------------------------------------------


def _first_group_member_email(group_name):
    sg = SystemGroup.objects.filter(name=group_name).first()
    if not sg:
        return None
    for uid in sg.get_system_group_member_ids():
        u = EmailUserRO.objects.filter(id=uid).first()
        if u and u.email:
            return u.email
    return None


assessor_email = _first_group_member_email(settings.GROUP_NAME_OCCURRENCE_ASSESSOR)
approver_email = _first_group_member_email(settings.GROUP_NAME_OCCURRENCE_APPROVER)
if not assessor_email or not approver_email:
    print("ERROR: Could not find assessor/approver group members.")
    sys.exit(1)
print(f"Using assessor email: {assessor_email}")
print(f"Using approver email: {approver_email}")

# ---------------------------------------------------------------------------
# 3. A real fauna species
# ---------------------------------------------------------------------------
fauna_species_name = (
    Taxonomy.objects.filter(species__group_type=fauna_gt).values_list("scientific_name", flat=True).first()
)
if not fauna_species_name:
    print("ERROR: No fauna species found in DB.")
    sys.exit(1)
print(f"Using fauna species: {fauna_species_name}")

# ---------------------------------------------------------------------------
# 4. A pre-existing Occurrence to use as the ORFAPP FK target in rows 7 & 9
# ---------------------------------------------------------------------------
existing_occ = (
    Occurrence.objects.filter(occurrence_number="OCC272230").first()
    or Occurrence.objects.filter(group_type=fauna_gt).order_by("-id").first()
)
if not existing_occ:
    print("ERROR: No existing fauna Occurrence found in DB for ORFAPP FK target.")
    sys.exit(1)
print(f"Using existing OCC for ORFAPP FK rows: {existing_occ.occurrence_number}")

# ---------------------------------------------------------------------------
# 5. Create (or clear) the test schema
# ---------------------------------------------------------------------------
SCHEMA_NAME = "OCC Link Test Schema"

schema = OccurrenceReportBulkImportSchema.objects.filter(name=SCHEMA_NAME, group_type=fauna_gt).first()
if schema:
    schema.columns.all().delete()
    print(f"Re-using existing schema: {schema} (id={schema.id}) -- columns cleared")
else:
    next_version = (
        OccurrenceReportBulkImportSchema.objects.filter(group_type=fauna_gt)
        .order_by("-version")
        .values_list("version", flat=True)
        .first()
        or 0
    ) + 1
    schema = OccurrenceReportBulkImportSchema.objects.create(
        name=SCHEMA_NAME,
        group_type=fauna_gt,
        version=next_version,
    )
    print(f"Created new schema: {schema} (id={schema.id})")

# ---------------------------------------------------------------------------
# 6. Define columns
# ---------------------------------------------------------------------------
ocr_ct = ContentType.objects.get_for_model(OccurrenceReport)
occ_ct = ContentType.objects.get_for_model(Occurrence)
orfapp_ct = ContentType.objects.get_for_model(OccurrenceReportApprovalDetails)

# Tuples: (header, field_name, content_type, allow_blank, lookup_field, is_emailuser)
COLUMNS = [
    # ORF Migrated From ID must be first (column 0) -- process_row reads row[0] directly
    ("ORF Migrated From ID", "migrated_from_id", ocr_ct, False, None, False),
    ("ORF Species", "species", ocr_ct, False, "taxonomy__scientific_name", False),
    ("ORF Processing Status", "processing_status", ocr_ct, False, None, False),
    ("ORF Assigned Officer", "assigned_officer", ocr_ct, True, None, True),
    ("ORF Assigned Approver", "assigned_approver", ocr_ct, True, None, True),
    ("ORF Approved By", "approved_by", ocr_ct, True, None, True),
    ("ORF Observation Date", "observation_date", ocr_ct, True, None, False),
    # OCC columns (migrated_from_id linking path)
    ("OCC Migrated From ID", "migrated_from_id", occ_ct, True, None, False),
    ("OCC Processing Status", "processing_status", occ_ct, True, None, False),
    ("OCC Comment", "comment", occ_ct, True, None, False),
    # ORFAPP columns (OccurrenceReportApprovalDetails linking paths)
    ("ORFAPP Occurrence", "occurrence", orfapp_ct, True, "occurrence_number", False),
    ("ORFAPP New Occurrence Name", "new_occurrence_name", orfapp_ct, True, None, False),
]

for header, field, ct, allow_blank, lookup, is_email in COLUMNS:
    col = OccurrenceReportBulkImportSchemaColumn(
        schema=schema,
        xlsx_column_header_name=header,
        django_import_content_type=ct,
        django_import_field_name=field,
        xlsx_data_validation_allow_blank=allow_blank,
        is_emailuser_column=is_email,
    )
    if lookup:
        col.django_lookup_field_name = lookup
    col.save()
    print(f"  Added column: {header}")

print(f"\nSchema id {schema.id} has {schema.columns.count()} columns.")

# ---------------------------------------------------------------------------
# 7. Build xlsx
# ---------------------------------------------------------------------------
OUT_PATH = os.path.join(settings.BASE_DIR, "occ_link_test.xlsx")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Import"

headers = list(schema.columns.order_by("order").values_list("xlsx_column_header_name", flat=True))
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)


def make_row(**kwargs):
    """Build a full row with empty-string defaults for all columns."""
    row = {h: "" for h in headers}
    row.update(kwargs)
    return [row[h] for h in headers]


# Shared kwargs for approved rows
_approved = {
    "ORF Assigned Officer": assessor_email,
    "ORF Assigned Approver": approver_email,
    "ORF Approved By": approver_email,
}

# --------------------------------------------------------------------------
# Row 2 -- OCC path: create OCC "occ-test-001", link approved ORF
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-001",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "01/01/2024",
            "OCC Migrated From ID": "occ-test-001",
            "OCC Processing Status": "active",
            "OCC Comment": "Created by row 2",
            **_approved,
        }
    )
)

# --------------------------------------------------------------------------
# Row 3 -- OCC path: second approved ORF links to same OCC (same migrated_from_id)
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-002",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "15/01/2024",
            "OCC Migrated From ID": "occ-test-001",
            "OCC Comment": "Linked by row 3",
            **_approved,
        }
    )
)

# --------------------------------------------------------------------------
# Row 4 -- OCC path: with_assessor ORF; OCC linked (occurrence FK set via reverse-FK add)
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-003",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "with_assessor",
            "ORF Observation Date": "20/01/2024",
            "OCC Migrated From ID": "occ-test-001",
            "OCC Comment": "Linked by row 4 (not approved)",
        }
    )
)

# --------------------------------------------------------------------------
# Row 5 -- OCC path: with_assessor ORF; no OCC fields -> no OCC created
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-004",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "with_assessor",
            "ORF Observation Date": "25/01/2024",
        }
    )
)

# --------------------------------------------------------------------------
# Row 6 -- ORFAPP path: approved ORF creates new OCC via ORFAPP New Occurrence Name
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-005",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "01/02/2024",
            "ORFAPP New Occurrence Name": "ORFAPP Created OCC",
            **_approved,
        }
    )
)

# --------------------------------------------------------------------------
# Row 7 -- ORFAPP path: approved ORF links to existing OCC via ORFAPP Occurrence FK
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-006",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "10/02/2024",
            "ORFAPP Occurrence": existing_occ.occurrence_number,
            **_approved,
        }
    )
)

# --------------------------------------------------------------------------
# Row 8 -- AMBIGUOUS: OCC Migrated From ID + ORFAPP New Occurrence Name -> error
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-007",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "15/02/2024",
            "OCC Migrated From ID": "occ-test-ambig-1",
            "ORFAPP New Occurrence Name": "Ambiguous OCC A",
            **_approved,
        }
    )
)

# --------------------------------------------------------------------------
# Row 9 -- AMBIGUOUS: OCC Migrated From ID + ORFAPP Occurrence FK -> error
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-008",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "20/02/2024",
            "OCC Migrated From ID": "occ-test-ambig-2",
            "ORFAPP Occurrence": existing_occ.occurrence_number,
            **_approved,
        }
    )
)

# --------------------------------------------------------------------------
# Row 10 -- AMBIGUOUS: ORFAPP Occurrence FK + ORFAPP New Occurrence Name -> error
# --------------------------------------------------------------------------
ws.append(
    make_row(
        **{
            "ORF Migrated From ID": "orf-009",
            "ORF Species": fauna_species_name,
            "ORF Processing Status": "approved",
            "ORF Observation Date": "25/02/2024",
            "ORFAPP Occurrence": existing_occ.occurrence_number,
            "ORFAPP New Occurrence Name": "Ambiguous OCC B",
            **_approved,
        }
    )
)

# Auto-width
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value))) + 4
for col, w in dims.items():
    ws.column_dimensions[get_column_letter(col)].width = w

wb.save(OUT_PATH)
print(f"\nTest xlsx written to: {OUT_PATH}")
print(f"Schema id to use when uploading: {schema.id}")
print(f"ORFAPP FK target OCC: {existing_occ.occurrence_number}")
print("""
Expected behaviour
------------------
Row  2: OCC path -- creates OCC (occ-test-001) + approved ORF linked via FK + M2M.
Row  3: OCC path -- links second approved ORF to same OCC.
Row  4: OCC path -- with_assessor ORF; OCC linked (occurrence FK set via reverse-FK add).
Row  5: No OCC fields -- no OCC created or linked.
Row  6: ORFAPP new_occurrence_name -- creates new OCC "ORFAPP Created OCC" + links approved ORF.
Row  7: ORFAPP occurrence FK -- links approved ORF to pre-existing OCC.
Row  8: ERROR -- ambiguous: OCC Migrated From ID + ORFAPP New Occurrence Name both set.
Row  9: ERROR -- ambiguous: OCC Migrated From ID + ORFAPP Occurrence FK both set.
Row 10: ERROR -- ambiguous: ORFAPP Occurrence FK + ORFAPP New Occurrence Name both set.
""")
